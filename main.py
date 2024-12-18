from flask import Flask, render_template, request, redirect, url_for
from openpyxl import load_workbook, Workbook

app = Flask(__name__)

# Define the route for the form
@app.route('/')
def form():
    return render_template('index.html')  # Create an HTML form template

# Route to handle form submission
@app.route('/submit', methods=['POST'])
def submit():
    # Get form data
    Name = request.form.get('Name')
    Rollno = request.form.get('Rollno')
    Age = request.form.get('Age')
    Branch = request.form.get('Branch')
    
    # Path to the Excel file
    excel_file = 'students.csv'
    
    try:
        # Load the workbook if it exists, otherwise create a new one
        try:
            workbook = load_workbook(excel_file)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            # Add header row if creating a new file
            sheet.append(["Name", "Rollno" , "Age","Branch"])
        
        # Add the data to the Excel file
        sheet.append([Name, Rollno , Age , Branch])
        
        # Save the workbook
        workbook.save(excel_file)
    except Exception as e:
        return f"An error occurred: {e}"
    
    return redirect(url_for('form'))  # Redirect back to the form page

if __name__ == '__main__':
    app.run(debug=True)